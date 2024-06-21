import os
import pandas as pd
from openpyxl import load_workbook
import openpyxl.utils.cell
import logging

# Setup logging
logging.basicConfig(filename='app.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def list_excel_files(directory):
    """List all Excel files in the given directory."""
    excel_extensions = ['.xls', '.xlsx']
    files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f)) and os.path.splitext(f)[1] in excel_extensions]
    logging.info(f"Excel files found: {files}")
    return files

def open_excel_file(filepath):
    """Open an Excel file."""
    try:
        return pd.ExcelFile(filepath)
    except Exception as e:
        logging.error(f"Failed to open {filepath}: {e}")
        return None

def read_config_sheet(filepath):
    """Read the config sheet to get mappings of sheet names to column names and cell numbers."""
    xl = open_excel_file(filepath)
    if xl:
        try:
            config_df = pd.read_excel(xl, sheet_name='config')
            return config_df
        except Exception as e:
            logging.error(f"Failed to read config sheet in {filepath}: {e}")
            return pd.DataFrame()  # Return empty DataFrame if there's an error
    return pd.DataFrame()

def append_row_to_sheet(filepath, sheet_name, config_df, row_data_dict):
    """Append a row to the selected sheet in the Excel file according to the config, maintaining formatting."""
    try:
        wb = load_workbook(filename=filepath)
        sheet = wb[sheet_name]

        # Filter relevant config for the selected sheet name
        relevant_config = config_df[config_df['Sheet Name'] == sheet_name]

        # Get the actual column letters from the relevant config rows
        column_mapping = {row[f'Cell Number {i//2 + 1}']: row[f'Column Name {i//2 + 1}']
                          for _, row in relevant_config.iterrows()
                          for i in range(1, len(config_df.columns), 2)
                          if not pd.isnull(row[f'Column Name {i//2 + 1}'])}

        # Determine the first completely empty row in the relevant columns
        last_row = max([find_last_row(sheet, ''.join(filter(str.isalpha, cell_ref))) for cell_ref in column_mapping.keys()], default=0)

        # Append data in each column specified in the config
        for cell_ref, column_name in column_mapping.items():
            col_letter = ''.join(filter(str.isalpha, cell_ref))
            col_idx = openpyxl.utils.cell.column_index_from_string(col_letter)
            cell_value = row_data_dict.get(column_name, None)  # Get the value based on column name

            if cell_value is not None:  # Only set the cell value if we have a value for this column
                target_cell = sheet.cell(row=last_row + 1, column=col_idx)

                # Check if the target cell is part of a merged cell range
                if any(target_cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges):
                    # Find the top-left cell of the merged range
                    for merged_cell_range in sheet.merged_cells.ranges:
                        if target_cell.coordinate in merged_cell_range:
                            top_left_cell = merged_cell_range.start_cell
                            sheet.cell(row=top_left_cell.row, column=top_left_cell.column, value=cell_value)
                            break
                else:
                    sheet.cell(row=last_row + 1, column=col_idx, value=cell_value)

        wb.save(filepath)
        logging.info(f"Appended new row to {sheet_name} in {filepath}")
    except Exception as e:
        logging.error(f"Error appending row to {sheet_name} in {filepath}: {e}")

def find_last_row(sheet, column):
    """Find the last row in the column that has data, considering only actual data, ignoring formulas."""
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet[f'{column}{row}'].value
        if cell_value is None or cell_value == '':
            return row - 1
    return sheet.max_row

def main():
    """Main function for handling Excel files."""
    directory = os.getcwd()  # Use the current working directory or adjust as needed
    excel_files = list_excel_files(directory)

    if excel_files:
        file_path = os.path.join(directory, excel_files[0])  # Example: Choose the first file
        config_df = read_config_sheet(file_path)
        if not config_df.empty:
            sheet_name = config_df.iloc[0]['Sheet Name']
            row_data_dict = {
                'Column Name 1': 'Value 1',
                'Column Name 2': 'Value 2',
                # Add more column data as needed
            }
            append_row_to_sheet(file_path, sheet_name, config_df, row_data_dict)
            print("Data appended successfully.")

# Uncomment and call main() if you are running this script outside of a Jupyter Notebook or Pyodide environment.
# main()



